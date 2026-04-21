#!/usr/bin/env python3
"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def setup_libreoffice_macro():
    """Setup LibreOffice macro for recalculation if not already configured."""
    system = platform.system()
    if system == "Darwin":
        macro_dir = os.path.expanduser(
            "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
        )
    elif system == "Windows":
        # Typical Windows path:
        #   %APPDATA%\LibreOffice\4\user\basic\Standard
        appdata = os.environ.get("APPDATA")
        if appdata:
            macro_dir = os.path.join(appdata, "LibreOffice", "4", "user", "basic", "Standard")
        else:
            macro_dir = os.path.expanduser("~\\AppData\\Roaming\\LibreOffice\\4\\user\\basic\\Standard")
    else:
        macro_dir = os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")

    macro_file = os.path.join(macro_dir, "Module1.xba")

    if os.path.exists(macro_file):
        with open(macro_file, "r", encoding="utf-8") as f:
            if "RecalculateAndSave" in f.read():
                return True

    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            check=False,
        )
        os.makedirs(macro_dir, exist_ok=True)

    macro_content = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

    try:
        with open(macro_file, "w", encoding="utf-8") as f:
            f.write(macro_content)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """Recalculate formulas in Excel file and report any errors."""
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    # Avoid timeout wrappers on Windows (upstream used `timeout`/`gtimeout` for unix).
    if platform.system() != "Windows":
        timeout_cmd = "timeout" if platform.system() == "Linux" else None
        if platform.system() == "Darwin":
            try:
                subprocess.run(
                    ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
                )
                timeout_cmd = "gtimeout"
            except (FileNotFoundError, subprocess.TimeoutExpired):
                pass

        if timeout_cmd:
            cmd = [timeout_cmd, str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True, check=False)

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result_obj = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result_obj["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result_obj["total_formulas"] = formula_count
        return result_obj

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(
            "Usage: python recalc.py <excel_file> [timeout_seconds]\n\n"
            "Recalculates all formulas via LibreOffice, then prints JSON with\n"
            "status, total_errors, error_summary, total_formulas.\n"
            "Optional timeout (seconds) wraps soffice on Linux/macOS only."
        )
        sys.exit(0 if len(sys.argv) >= 2 and sys.argv[1] in ("-h", "--help") else 1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()

